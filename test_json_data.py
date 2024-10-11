
# import os
# print(os.getcwd())
import json
import sys
sys.path.append('.')
from  core.models import Flowline, ProjectDataModel,Names


with open('./tests/proj_test.json', 'r') as json_file:
    data_json = json.load(json_file)
    proj_model= ProjectDataModel(**data_json)

def test_init_correct_lenData_of_field():
    sut=proj_model

    assert len(sut.FlowlineData)>0
    assert len(sut.Names)>0
    assert len(sut.Network)>0
    assert len(sut.InData)>0
    assert len(sut.OutData)==1

def test_get_typesofNames():
    sut=Names.get_types()
    assert type(sut) is tuple 
    assert len(sut) > 0 

def test_get_trace():
    sut=proj_model.get_trace("kg205.Flowline_1")
    # print(list(proj_model.get_trace("kg202.Flowline_1",True).values())[0].__dict__)
    # print(proj_model.get_trace("kg205.Flowline_1"))
    # print(proj_model.get_trace("kg202-205.Flowline_1"))



if __name__ == '__main__':
    import pandas as pd
    from openpyxl import load_workbook
    # for k,v in data_json.items():
    #     print(k,type(v))
    # print(len(proj_model.Network))
    # for n in proj_model.Network:
    #     print(n.__dict__)
    # print(proj_model.to_json(True))

    # print(proj_model.to_json(True))
    
    # test_get_trace()
    # print(proj_model.FlowlineData)

    # net=proj_model.Network
    # find_str='Ck'
    # for b in net:
    #     if b.Destination==find_str or b.Source==find_str:
    #         print(b)
    # print(proj_model.to_json(True))
    # print(data_json)
    names={}
    for name in proj_model.Names.Flowline:
        names=dict(list(names.items()) + list(proj_model.get_trace(name).items())) 
    p=pd.DataFrame.from_dict(proj_model.FlowlineData,orient='index')
    n=pd.DataFrame.from_dict(names,orient='index')
    # pof=pd.DataFrame.from_dict(proj_model.FlowlineData,orient='index')
    df=n.join(p)
    # df['Name']=df.index
    df=df.drop('GeoProfile', axis=1)
    print(df)
    book = load_workbook('test.xlsx')
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    writer.book = book
    # df.to_excel('test.xlsx', index_label="Name")
    df.to_excel(writer,index_label="Name",sheet_name='Sheet1', index=False, startrow=book['Sheet1'].max_row)
    df.to_excel(writer,index_label="Name",sheet_name='Sheet1', index=False, startrow=book['Sheet1'].max_row+5)
    writer.save()
    # print(n)
    pass
